import streamlit as st
import pandas as pd
from utils.db_utils import get_connection
from io import BytesIO

# ==========================
# 🎨 CSS Styling cho Admin Panel với thông báo đóng được
# ==========================
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global Styling */
    .stApp {
        background: linear-gradient(135deg, #f0f8ff 0%, #e6f3ff 100%);
        font-family: 'Inter', sans-serif;
    }
    
    /* Header Admin */
    .admin-header {
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 50%, #0f172a 100%);
        padding: 2.5rem 2rem;
        border-radius: 20px;
        margin-bottom: 2rem;
        box-shadow: 0 15px 35px rgba(59, 130, 246, 0.3);
        text-align: center;
        position: relative;
        overflow: hidden;
    }
    
    .admin-header::before {
        content: '⚙️';
        position: absolute;
        top: 20px;
        right: 30px;
        font-size: 3rem;
        opacity: 0.1;
    }
    
    .admin-header h1 {
        color: white;
        font-size: 2.8rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 0 2px 8px rgba(0,0,0,0.3);
    }
    
    .admin-header .subtitle {
        color: rgba(255, 255, 255, 0.9);
        font-size: 1.2rem;
        margin-top: 0.8rem;
        font-weight: 400;
    }
    
    /* Tab Styling - Fixed để hiển thị đầy đủ */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: white;
        border-radius: 15px;
        padding: 0.5rem;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.1);
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        border: 2px solid transparent !important;
        border-radius: 12px !important;
        padding: 0.8rem 1.5rem !important;
        font-weight: 600 !important;
        color: #1e40af !important;
        transition: all 0.3s ease !important;
        white-space: nowrap !important;
        min-width: fit-content !important;
    }
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, #3b82f6, #1e40af) !important;
        color: white !important;
        border: 2px solid #1e40af !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3) !important;
    }
    
    .stTabs [data-baseweb="tab"]:hover:not([aria-selected="true"]) {
        background: rgba(59, 130, 246, 0.1) !important;
        border: 2px solid #3b82f6 !important;
    }
    
    /* Card Container */
    .admin-card {
        background: white;
        padding: 2rem;
        border-radius: 18px;
        box-shadow: 0 8px 32px rgba(59, 130, 246, 0.12);
        border: 2px solid #e0f2fe;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
    }
    
    .admin-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 5px;
        background: linear-gradient(90deg, #3b82f6, #60a5fa, #93c5fd);
    }
    
    .admin-card h3 {
        color: #1e40af;
        font-weight: 700;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        gap: 0.8rem;
        font-size: 1.4rem;
    }
    
    /* Statistics Cards */
    .stat-card {
        background: linear-gradient(135deg, #dbeafe, #bfdbfe);
        padding: 1.5rem;
        border-radius: 12px;
        text-align: center;
        border: 1px solid #93c5fd;
        box-shadow: 0 4px 12px rgba(147, 197, 253, 0.2);
        transition: transform 0.3s ease;
    }
    
    .stat-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(147, 197, 253, 0.3);
    }
    
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        color: #1e40af;
        margin-bottom: 0.5rem;
    }
    
    .stat-label {
        color: #1e40af;
        font-weight: 500;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* Filter Section */
    .filter-section {
        background: #f8fafc;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #e2e8f0;
        margin-bottom: 1.5rem;
    }
    
    .filter-title {
        color: #1e40af;
        font-weight: 600;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    /* Button Styling */
    .stButton > button {
        background: linear-gradient(135deg, #3b82f6, #1e40af) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.7rem 1.5rem !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3) !important;
        transition: all 0.3s ease !important;
        width: 100% !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(59, 130, 246, 0.4) !important;
    }
    
    /* Download Button Styling */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #059669, #10b981) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.7rem 1.5rem !important;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.3) !important;
        transition: all 0.3s ease !important;
        width: 100% !important;
        margin-bottom: 0.5rem !important;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4) !important;
    }
    
    /* Update Button */
    .update-btn button {
        background: linear-gradient(135deg, #f59e0b, #d97706) !important;
    }
    
    .update-btn button:hover {
        box-shadow: 0 6px 20px rgba(245, 158, 11, 0.4) !important;
    }
    
    /* Lock Button */
    .lock-btn button {
        background: linear-gradient(135deg, #dc2626, #b91c1c) !important;
    }
    
    .lock-btn button:hover {
        box-shadow: 0 6px 20px rgba(220, 38, 38, 0.4) !important;
    }
    
    /* Delete Button */
    .delete-btn button {
        background: linear-gradient(135deg, #991b1b, #7f1d1d) !important;
    }
    
    .delete-btn button:hover {
        box-shadow: 0 6px 20px rgba(153, 27, 27, 0.5) !important;
    }
    
    /* Upload Section */
    .upload-section {
        background: linear-gradient(135deg, #f0f9ff, #e0f2fe);
        padding: 2rem;
        border-radius: 15px;
        border: 2px dashed #60a5fa;
        text-align: center;
        margin: 1.5rem 0;
    }
    
    .upload-icon {
        font-size: 3rem;
        margin-bottom: 1rem;
        color: #3b82f6;
    }
    
    /* DataFrames */
    .stDataFrame {
        border: 1px solid #e0f2fe !important;
        border-radius: 12px !important;
        overflow: hidden !important;
    }
    
    /* Alert Styling - Updated for closeable notifications */
    .stAlert {
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1) !important;
        margin: 1rem 0 !important;
    }
    
    /* Closeable Notification Styles */
    .closeable-notification {
        position: relative;
        padding: 1rem 1.5rem 1rem 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        animation: slideInDown 0.3s ease-out;
        border-left: 4px solid;
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
    }
    
    @keyframes slideInDown {
        from {
            opacity: 0;
            transform: translateY(-10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .notification-content {
        flex: 1;
        margin-right: 1rem;
    }
    
    .notification-close {
        background: none;
        border: none;
        font-size: 1.2rem;
        cursor: pointer;
        padding: 0;
        color: inherit;
        opacity: 0.7;
        transition: opacity 0.2s ease;
        width: 20px;
        height: 20px;
        display: flex;
        align-items: center;
        justify-content: center;
        border-radius: 50%;
        flex-shrink: 0;
    }
    
    .notification-close:hover {
        opacity: 1;
        background: rgba(0,0,0,0.1);
    }
    
    /* Success Notification */
    .notification-success {
        background: linear-gradient(135deg, #d1fae5, #a7f3d0);
        color: #065f46;
        border-left-color: #10b981;
    }
    
    /* Error Notification */
    .notification-error {
        background: linear-gradient(135deg, #fee2e2, #fecaca);
        color: #991b1b;
        border-left-color: #ef4444;
    }
    
    /* Warning Notification */
    .notification-warning {
        background: linear-gradient(135deg, #fef3c7, #fde68a);
        color: #92400e;
        border-left-color: #f59e0b;
    }
    
    /* Info Notification */
    .notification-info {
        background: linear-gradient(135deg, #dbeafe, #bfdbfe);
        color: #1e40af;
        border-left-color: #3b82f6;
    }
    
    /* Role Badge */
    .role-badge {
        display: inline-block;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .role-admin { background: linear-gradient(135deg, #dc2626, #ef4444); color: white; }
    .role-doctor { background: linear-gradient(135deg, #059669, #10b981); color: white; }
    .role-user { background: linear-gradient(135deg, #3b82f6, #60a5fa); color: white; }
    
    /* Status Badge */
    .status-badge {
        display: inline-block;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .status-active { background: linear-gradient(135deg, #059669, #10b981); color: white; }
    .status-locked { background: linear-gradient(135deg, #dc2626, #ef4444); color: white; }
    
    /* User Management Cards */
    .user-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #e2e8f0;
        margin-bottom: 1rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        transition: all 0.3s ease;
    }
    
    .user-card:hover {
        box-shadow: 0 4px 16px rgba(59, 130, 246, 0.1);
        border-color: #93c5fd;
    }
    
    .user-info {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 1rem;
    }
    
    .user-details h4 {
        color: #1e40af;
        margin: 0 0 0.5rem 0;
        font-size: 1.1rem;
    }
    
    .user-details p {
        color: #64748b;
        margin: 0;
        font-size: 0.9rem;
    }
    
    .user-actions {
        display: flex;
        gap: 0.5rem;
        flex-wrap: wrap;
    }
    
    /* Export Section */
    .export-section {
        background: #f8fafc;
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid #e2e8f0;
        margin-top: 2rem;
    }
    
    .export-title {
        color: #1e40af;
        font-weight: 600;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-size: 1.2rem;
    }
    
    /* Responsive */
    @media (max-width: 768px) {
        .admin-header h1 {
            font-size: 2rem;
        }
        .admin-card {
            padding: 1.5rem;
        }
        .user-info {
            flex-direction: column;
            align-items: flex-start;
        }
        .user-actions {
            margin-top: 1rem;
            width: 100%;
        }
        .closeable-notification {
            flex-direction: column;
        }
        .notification-close {
            align-self: flex-end;
            margin-top: 0.5rem;
        }
    }
    
    /* Hide notifications that are dismissed */
    .notification-hidden {
        display: none !important;
    }
</style>

<script>
// JavaScript for closing notifications
function closeNotification(notificationId) {
    const notification = document.getElementById(notificationId);
    if (notification) {
        notification.style.animation = 'slideOutUp 0.3s ease-in';
        setTimeout(() => {
            notification.classList.add('notification-hidden');
        }, 300);
    }
}

// CSS animation for closing
const style = document.createElement('style');
style.textContent = `
@keyframes slideOutUp {
    from {
        opacity: 1;
        transform: translateY(0);
    }
    to {
        opacity: 0;
        transform: translateY(-20px);
    }
}
`;
document.head.appendChild(style);
</script>
""", unsafe_allow_html=True)

st.set_page_config(
    page_title="Admin Dashboard", 
    layout="wide",
    page_icon="🛠️",
    initial_sidebar_state="collapsed"
)
# ==========================
# Notification Functions
# ==========================
def show_notification(message, type="info", auto_close=False):
    """
    Hiển thị thông báo có thể đóng
    type: success, error, warning, info
    """
    notification_id = f"notification_{type}_{hash(message) % 10000}"
    
    # Map notification types to icons
    icons = {
        "success": "✅",
        "error": "❌", 
        "warning": "⚠️",
        "info": "ℹ️"
    }
    
    auto_close_script = ""
    if auto_close:
        auto_close_script = f"""
        setTimeout(() => {{
            closeNotification('{notification_id}');
        }}, 5000); // Auto close after 5 seconds
        """
    
    st.markdown(f"""
    <div id="{notification_id}" class="closeable-notification notification-{type}">
        <div class="notification-content">
            <strong>{icons.get(type, "ℹ️")} {message}</strong>
        </div>
        <button class="notification-close" onclick="closeNotification('{notification_id}')" title="Đóng thông báo">
            ×
        </button>
    </div>
    
    <script>
    {auto_close_script}
    </script>
    """, unsafe_allow_html=True)

# ==========================
# Initialize session state for notifications
# ==========================
if 'notifications' not in st.session_state:
    st.session_state.notifications = []

# ==========================
# 🎨 Header Section
# ==========================
st.markdown("""
<div class="admin-header">
    <h1>🛠️ Admin Dashboard</h1>
    <div class="subtitle">Hệ thống quản trị toàn diện - AI Healthcare Platform</div>
</div>
""", unsafe_allow_html=True)

# ================== CHECK QUYỀN ==================
if "user" not in st.session_state or not st.session_state.logged_in:
    show_notification("Yêu cầu đăng nhập - Vui lòng đăng nhập bằng tài khoản quản trị viên để truy cập trang này.", "error")
    st.stop()

if st.session_state.user.get("role") != "admin":
    show_notification("Không có quyền truy cập - Bạn không có quyền truy cập vào trang quản trị này.", "error")
    st.stop()

# ==========================
# Welcome Message
# ==========================
show_notification(f"Chào mừng quản trị viên: {st.session_state.user.get('username', 'Admin')} - Bạn đang truy cập với quyền cao nhất trong hệ thống.", "info")

# ================== TAB MENU ==================
tab1, tab2, tab3 = st.tabs([
    "📊 Dashboard & Lịch sử", 
    "👥 Quản lý người dùng", 
    "📋 Logs & Báo cáo"
])

# ================== TAB 1: DASHBOARD & LỊCH SỬ ==================
with tab1:
    # Stats Overview
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Lấy thống kê tổng quan
        cursor.execute("SELECT COUNT(*) as total_users FROM users")
        total_users = cursor.fetchone()['total_users']
        
        cursor.execute("SELECT COUNT(*) as total_diagnoses FROM lich_su_chan_doan")
        total_diagnoses = cursor.fetchone()['total_diagnoses']
        
        cursor.execute("SELECT COUNT(DISTINCT algorithm) as total_algorithms FROM lich_su_chan_doan")
        total_algorithms = cursor.fetchone()['total_algorithms']
        
        cursor.execute("SELECT COUNT(*) as total_logs FROM activity_logs")
        total_logs = cursor.fetchone()['total_logs']
        
        # Hiển thị stats
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-number">{total_users}</div>
                <div class="stat-label">👥 Người dùng</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col2:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-number">{total_diagnoses}</div>
                <div class="stat-label">🏥 Chẩn đoán</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col3:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-number">{total_algorithms}</div>
                <div class="stat-label">🤖 AI Models</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col4:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-number">{total_logs}</div>
                <div class="stat-label">📋 Activity Logs</div>
            </div>
            """, unsafe_allow_html=True)
        
    except Exception as e:
        show_notification(f"Lỗi khi lấy thống kê: {e}", "error")
    
    # Lịch sử chẩn đoán
    st.markdown('<div class="admin-card">', unsafe_allow_html=True)
    st.markdown('<h3>📊 Lịch sử chẩn đoán toàn hệ thống</h3>', unsafe_allow_html=True)
    
    try:
        cursor.execute("""
            SELECT u.username, l.*, 
                   DATE_FORMAT(l.created_at, '%d/%m/%Y %H:%i') as formatted_date
            FROM lich_su_chan_doan l
            JOIN users u ON l.user_id = u.id
            ORDER BY l.created_at DESC
        """)
        rows = cursor.fetchall()
        cursor.close(); conn.close()

        if rows:
                df = pd.DataFrame(rows)
    
        # Format độ tin cậy nếu có cột này
        if 'confidence' in df.columns:
            def format_confidence(val):
                if pd.isna(val) or val is None:
                    return "N/A"
                try:
                    confidence = float(val)
                    if confidence <= 1:   # nếu <= 1 thì coi là tỷ lệ (0.98 -> 98%)
                        confidence *= 100
                    confidence = round(confidence, 1)
                    # Nếu là số nguyên thì bỏ .0
                    if confidence.is_integer():
                        return f"{int(confidence)}%"
                    return f"{confidence}%"
                except (ValueError, TypeError):
                    return str(val)

            df['confidence'] = df['confidence'].apply(format_confidence)
            # Filter Section
            st.markdown("""
            <div class="filter-section">
                <div class="filter-title">🔍 Bộ lọc dữ liệu</div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            user_filter = col1.selectbox("👤 Người dùng", ["Tất cả"] + sorted(df["username"].unique()))
            algo_filter = col2.selectbox("🤖 Thuật toán", ["Tất cả"] + sorted(df["algorithm"].unique()))
            severity_filter = col3.selectbox("⚕️ Mức độ", ["Tất cả"] + sorted(df["severity"].unique()) if "severity" in df.columns else ["Tất cả"])

            # Apply filters
            filtered_df = df.copy()
            if user_filter != "Tất cả":
                filtered_df = filtered_df[filtered_df["username"] == user_filter]
            if algo_filter != "Tất cả":
                filtered_df = filtered_df[filtered_df["algorithm"] == algo_filter]
            if severity_filter != "Tất cả" and "severity" in df.columns:
                filtered_df = filtered_df[filtered_df["severity"] == severity_filter]

            # Display data
            st.dataframe(
                filtered_df.drop('created_at', axis=1) if 'created_at' in filtered_df.columns else filtered_df,
                use_container_width=True, 
                height=500
            )
        else:
            show_notification("Chưa có dữ liệu chẩn đoán nào trong hệ thống.", "info")
    except Exception as e:
        show_notification(f"Lỗi khi tải dữ liệu: {e}", "error")
    
    st.markdown('</div>', unsafe_allow_html=True)

# ================== TAB 2: QUẢN LÝ USER NÂNG CAO ==================
with tab2:
    st.markdown('<div class="admin-card">', unsafe_allow_html=True)
    st.markdown('<h3>👥 Quản lý người dùng nâng cao</h3>', unsafe_allow_html=True)

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # Thêm cột status nếu chưa có
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN status ENUM('active', 'locked') DEFAULT 'active'")
            conn.commit()
        except:
            pass

        # Lấy danh sách user
        cursor.execute("""
            SELECT id, username, email, role, status,
                   DATE_FORMAT(created_at, '%d/%m/%Y %H:%i') as join_date
            FROM users 
            ORDER BY created_at DESC
        """)
        users = cursor.fetchall()

        if users:
            # Hiển thị bảng user
            df_users = pd.DataFrame(users)
            st.dataframe(df_users, use_container_width=True, height=400)

            st.markdown("---")
            st.markdown("### ⚙️ Công cụ quản lý người dùng")

            # Chọn user
            col1, col2 = st.columns([2, 1])
            with col1:
                selected_user = st.selectbox(
                    "👤 Chọn người dùng để quản lý:",
                    options=[(u['id'], f"{u['username']} ({u['email']}) - {u['role'].upper()}") for u in users],
                    format_func=lambda x: x[1]
                )
                user_id = selected_user[0]
                selected_user_info = next(u for u in users if u['id'] == user_id)

            with col2:
                st.markdown(f"""
                <div class="user-card">
                    <div class="user-details">
                        <h4>👤 {selected_user_info['username']}</h4>
                        <p>📧 {selected_user_info['email']}</p>
                        <p>🎭 Role: <span class="role-badge role-{selected_user_info['role']}">{selected_user_info['role'].upper()}</span></p>
                        <p>📅 Tham gia: {selected_user_info['join_date']}</p>
                        <p>🔒 Trạng thái: <span class="status-badge status-{selected_user_info['status']}">{selected_user_info['status'].upper()}</span></p>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("---")

            # Các hành động quản lý
            col1, col2, col3, col4 = st.columns(4)

            # 🎭 Phân quyền
            with col1:
                st.markdown("#### 🎭 Phân quyền")
                new_role = st.selectbox("Chọn quyền mới:", ["user", "admin"],
                                        index=["user", "admin"].index(selected_user_info['role']))
                if st.button("🔄 Cập nhật quyền", key="update_role", use_container_width=True):
                    if new_role != selected_user_info['role']:
                        try:
                            cursor.execute("UPDATE users SET role=%s WHERE id=%s", (new_role, user_id))
                            conn.commit()
                            show_notification("Cập nhật quyền thành công!", "success", auto_close=True)
                            st.rerun()
                        except Exception as e:
                            show_notification(f"Lỗi khi cập nhật: {e}", "error")
                    else:
                        show_notification("Quyền không thay đổi!", "warning", auto_close=True)

            # 🔒 Khóa / mở khóa
            with col2:
                st.markdown("#### 🔒 Trạng thái tài khoản")
                if selected_user_info['status'] == 'active':
                    if st.button("🔒 Khóa tài khoản", key="lock_user", use_container_width=True):
                        try:
                            cursor.execute("UPDATE users SET status='locked' WHERE id=%s", (user_id,))
                            conn.commit()
                            cursor.execute("""
                                INSERT INTO activity_logs (user_id, action, details, ip_address) 
                                VALUES (%s, %s, %s, %s)
                            """, (st.session_state.user.get('id'), 'LOCK_USER',
                                  f'Khóa tài khoản: {selected_user_info["username"]}', 'admin'))
                            conn.commit()
                            show_notification("Đã khóa tài khoản thành công!", "success", auto_close=True)
                            st.rerun()
                        except Exception as e:
                            show_notification(f"Lỗi khi khóa: {e}", "error")
                else:
                    if st.button("🔓 Mở khóa tài khoản", key="unlock_user", use_container_width=True):
                        try:
                            cursor.execute("UPDATE users SET status='active' WHERE id=%s", (user_id,))
                            conn.commit()
                            cursor.execute("""
                                INSERT INTO activity_logs (user_id, action, details, ip_address) 
                                VALUES (%s, %s, %s, %s)
                            """, (st.session_state.user.get('id'), 'UNLOCK_USER',
                                  f'Mở khóa tài khoản: {selected_user_info["username"]}', 'admin'))
                            conn.commit()
                            show_notification("Đã mở khóa tài khoản thành công!", "success", auto_close=True)
                            st.rerun()
                        except Exception as e:
                            show_notification(f"Lỗi khi mở khóa: {e}", "error")

            # 🔑 Reset password
            with col3:
                st.markdown("#### 🔑 Đặt lại mật khẩu")
                if st.button("🔑 Reset Password", key="reset_password", use_container_width=True):
                    try:
                        import hashlib
                        temp_password = "TempPass123!"
                        hashed_password = hashlib.sha256(temp_password.encode()).hexdigest()
                        cursor.execute("UPDATE users SET password=%s WHERE id=%s", (hashed_password, user_id))
                        conn.commit()
                        cursor.execute("""
                            INSERT INTO activity_logs (user_id, action, details, ip_address) 
                            VALUES (%s, %s, %s, %s)
                        """, (st.session_state.user.get('id'), 'RESET_PASSWORD',
                              f'Reset password cho: {selected_user_info["username"]}', 'admin'))
                        conn.commit()
                        show_notification(f"Đã reset password thành công! Mật khẩu tạm thời: {temp_password}", "success")
                    except Exception as e:
                        show_notification(f"Lỗi khi reset password: {e}", "error")

            # 🗑️ Xóa user
            with col4:
                st.markdown("#### ⚠️ Xóa tài khoản")
                cursor.execute("SELECT COUNT(*) as admin_count FROM users WHERE role='admin' AND status='active'")
                admin_count = cursor.fetchone()['admin_count']
                if selected_user_info['role'] == 'admin' and admin_count <= 1:
                    show_notification("Không thể xóa admin cuối cùng!", "warning")
                    st.button("🚫 Không thể xóa", disabled=True, use_container_width=True)
                else:
                    confirm_delete = st.checkbox(f"✅ Tôi xác nhận xóa user: {selected_user_info['username']}", key="confirm_delete")
                    if st.button("🗑️ XÓA VĨNH VIỄN", key="delete_user", use_container_width=True, disabled=not confirm_delete):
                        try:
                            cursor.execute("DELETE FROM lich_su_chan_doan WHERE user_id=%s", (user_id,))
                            cursor.execute("DELETE FROM activity_logs WHERE user_id=%s", (user_id,))
                            cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
                            conn.commit()
                            cursor.execute("""
                                INSERT INTO activity_logs (user_id, action, details, ip_address) 
                                VALUES (%s, %s, %s, %s)
                            """, (st.session_state.user.get('id'), 'DELETE_USER',
                                  f'Đã xóa tài khoản: {selected_user_info["username"]} - {selected_user_info["email"]}', 'admin'))
                            conn.commit()
                            show_notification("Đã xóa tài khoản và toàn bộ dữ liệu liên quan!", "success", auto_close=True)
                            st.rerun()
                        except Exception as e:
                            show_notification(f"Lỗi khi xóa tài khoản: {e}", "error")

            # 📊 Thống kê
            st.markdown("---")
            st.markdown("### 📊 Thống kê chi tiết người dùng")

            col1, col2, col3 = st.columns(3)
            cursor.execute("SELECT COUNT(*) as diagnosis_count FROM lich_su_chan_doan WHERE user_id=%s", (user_id,))
            diagnosis_count = cursor.fetchone()['diagnosis_count']

            cursor.execute("""
                SELECT DATE_FORMAT(MAX(created_at), '%d/%m/%Y %H:%i') as last_activity
                FROM activity_logs WHERE user_id=%s
            """, (user_id,))
            last_activity_result = cursor.fetchone()
            last_activity = last_activity_result['last_activity'] if last_activity_result and last_activity_result['last_activity'] else "Chưa có hoạt động"

            cursor.execute("""
                SELECT algorithm, COUNT(*) as count 
                FROM lich_su_chan_doan WHERE user_id=%s 
                GROUP BY algorithm ORDER BY count DESC LIMIT 1
            """, (user_id,))
            fav_algo_result = cursor.fetchone()
            fav_algo = f"{fav_algo_result['algorithm']} ({fav_algo_result['count']} lần)" if fav_algo_result else "Chưa có"

            with col1:
                st.markdown(f"""<div class="stat-card"><div class="stat-number">{diagnosis_count}</div><div class="stat-label">🏥 Số lần chẩn đoán</div></div>""", unsafe_allow_html=True)
            with col2:
                st.markdown(f"""<div class="stat-card"><div class="stat-number" style="font-size: 1.2rem;">{last_activity}</div><div class="stat-label">⏰ Hoạt động cuối</div></div>""", unsafe_allow_html=True)
            with col3:
                st.markdown(f"""<div class="stat-card"><div class="stat-number" style="font-size: 1rem;">{fav_algo}</div><div class="stat-label">🤖 Thuật toán yêu thích</div></div>""", unsafe_allow_html=True)

            # 📋 Lịch sử hoạt động
            if st.expander(f"📋 Lịch sử hoạt động của {selected_user_info['username']}", expanded=False):
                cursor.execute("""
                    SELECT l.*, DATE_FORMAT(l.created_at, '%d/%m/%Y %H:%i') as formatted_date
                    FROM lich_su_chan_doan l 
                    WHERE l.user_id=%s 
                    ORDER BY l.created_at DESC
                    LIMIT 20
                """, (user_id,))
                user_history = cursor.fetchall()

                # Format độ tin cậy cho user history
            if user_history:
                for row in user_history:
                    if 'confidence' in row and row['confidence'] is not None:
                        try:
                            val = float(row['confidence'])
                            # Nếu giá trị <= 1 thì đổi sang %
                            if val <= 1:
                                val *= 100
                            row['confidence'] = f"{val:.1f}%"
                        except (ValueError, TypeError):
                            row['confidence'] = str(row['confidence'])

                if user_history:
                    df_history = pd.DataFrame(user_history)
                    if 'created_at' in df_history.columns:
                        df_history = df_history.drop('created_at', axis=1)
                    st.dataframe(df_history, use_container_width=True, height=300)
                else:
                    show_notification("Người dùng này chưa có hoạt động chẩn đoán nào.", "info")

        else:
            show_notification("Chưa có người dùng nào trong hệ thống.", "info")

        cursor.close()
        conn.close()

    except Exception as e:
        show_notification(f"Lỗi khi tải dữ liệu người dùng: {e}", "error")
    st.markdown('</div>', unsafe_allow_html=True)

# ================== TAB 3: LOG & EXPORT ==================
with tab3:
    st.markdown('<div class="admin-card">', unsafe_allow_html=True)
    st.markdown('<h3>📋 Activity Logs & Báo cáo hệ thống</h3>', unsafe_allow_html=True)

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT *, 
                   DATE_FORMAT(created_at, '%d/%m/%Y %H:%i:%s') as formatted_date
            FROM activity_logs 
            ORDER BY created_at DESC 
            LIMIT 500
        """)
        logs = cursor.fetchall()
        cursor.close(); conn.close()

        if logs:
            df_logs = pd.DataFrame(logs)
            
            # Log filters
            st.markdown("""
            <div class="filter-section">
                <div class="filter-title">🔍 Bộ lọc logs</div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            if 'action' in df_logs.columns:
                action_filter = col1.selectbox("📝 Loại hoạt động:", ["Tất cả"] + sorted(df_logs["action"].unique()))
                if action_filter != "Tất cả":
                    df_logs = df_logs[df_logs["action"] == action_filter]
            
            if 'user_id' in df_logs.columns:
                user_filter = col2.selectbox("👤 Người dùng (ID):", ["Tất cả"] + sorted(df_logs["user_id"].unique().astype(str)))
                if user_filter != "Tất cả":
                    df_logs = df_logs[df_logs["user_id"].astype(str) == user_filter]
            
            # Display logs
            st.dataframe(
                df_logs.drop('created_at', axis=1) if 'created_at' in df_logs.columns else df_logs,
                use_container_width=True, 
                height=400
            )

            # Export Section - DROPDOWN VERSION
            st.markdown("""
            <div class="export-section">
                <div class="export-title">📤 Xuất dữ liệu báo cáo</div>
                <p style="color: #64748b; margin-bottom: 1.5rem;">Chọn định dạng để tải xuống dữ liệu với hỗ trợ đầy đủ tiếng Việt</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Dropdown để chọn định dạng xuất
            export_format = st.selectbox(
                "🎯 Chọn định dạng xuất file:",
                options=["CSV - Excel tự nhận diện tiếng Việt", 
                         "Excel - Định dạng đẹp với thống kê", 
                         "JSON - Giữ nguyên ký tự tiếng Việt", 
                         "PDF - Layout chuyên nghiệp"],
                help="Chọn định dạng phù hợp cho nhu cầu sử dụng của bạn"
            )

            # Nút xuất duy nhất
            if st.button("⬇️ Tải xuống", type="primary", use_container_width=True):
                try:
                    if "CSV" in export_format:
                        # CSV Export
                        import io
                        csv_buffer = io.StringIO()
                        df_logs.to_csv(csv_buffer, index=False, encoding='utf-8')
                        csv_content = csv_buffer.getvalue()
                        csv_with_bom = '\ufeff' + csv_content
                        csv_bytes = csv_with_bom.encode('utf-8')
                        
                        st.download_button(
                            "📊 Tải file CSV",
                            csv_bytes,
                            f"activity_logs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                            "text/csv",
                            help="CSV với UTF-8 BOM - Excel sẽ tự nhận diện tiếng Việt",
                            use_container_width=True
                        )
                        show_notification("File CSV đã sẵn sàng tải xuống!", "success", auto_close=True)

                    elif "Excel" in export_format:
                        # Excel Export
                        from io import BytesIO
                        output = BytesIO()
                        
                        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                            # Xuất dữ liệu chính
                            df_logs.to_excel(writer, index=False, sheet_name="Activity_Logs")
                            
                            # Tạo sheet thống kê
                            summary_data = {
                                'Thống kê': [
                                    'Tổng số logs',
                                    'Số loại hoạt động',
                                    'Số người dùng',
                                    'Thời gian đầu tiên',
                                    'Thời gian cuối cùng'
                                ],
                                'Giá trị': [
                                    len(df_logs),
                                    df_logs['action'].nunique() if 'action' in df_logs.columns else 0,
                                    df_logs['user_id'].nunique() if 'user_id' in df_logs.columns else 0,
                                    df_logs['formatted_date'].min() if 'formatted_date' in df_logs.columns else 'N/A',
                                    df_logs['formatted_date'].max() if 'formatted_date' in df_logs.columns else 'N/A'
                                ]
                            }
                            summary_df = pd.DataFrame(summary_data)
                            summary_df.to_excel(writer, sheet_name="Thống_kê", index=False)
                            
                            # Format Excel
                            try:
                                from openpyxl.styles import Font, PatternFill, Alignment
                                workbook = writer.book
                                
                                # Format main sheet
                                worksheet = writer.sheets["Activity_Logs"]
                                for cell in worksheet[1]:  # Header row
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center")
                                
                                # Auto-adjust column width
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max(max_length + 2, 10), 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                                
                                # Format summary sheet
                                summary_ws = writer.sheets["Thống_kê"]
                                for cell in summary_ws[1]:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                                summary_ws.column_dimensions['A'].width = 25
                                summary_ws.column_dimensions['B'].width = 20
                            except ImportError:
                                pass
                        
                        st.download_button(
                            "📈 Tải file Excel",
                            output.getvalue(),
                            f"activity_logs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Excel với định dạng đẹp và sheet thống kê",
                            use_container_width=True
                        )
                        show_notification("File Excel đã sẵn sàng tải xuống!", "success", auto_close=True)

                    elif "JSON" in export_format:
                        # JSON Export
                        import json
                        
                        df_json = df_logs.copy()
                        
                        # Chuyển đổi datetime/timestamp columns thành string
                        for col in df_json.columns:
                            if df_json[col].dtype == 'datetime64[ns]' or 'datetime' in str(df_json[col].dtype):
                                df_json[col] = df_json[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                            elif any(isinstance(val, pd.Timestamp) for val in df_json[col].dropna()):
                                df_json[col] = df_json[col].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and hasattr(x, 'strftime') else str(x))
                        
                        json_dict = df_json.to_dict(orient='records')
                        json_string = json.dumps(json_dict, indent=2, ensure_ascii=False)
                        json_data = json_string.encode('utf-8')
                        
                        st.download_button(
                            "📋 Tải file JSON",
                            json_data,
                            f"activity_logs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.json",
                            "application/json",
                            help="JSON với UTF-8 encoding, giữ nguyên tiếng Việt",
                            use_container_width=True
                        )
                        show_notification("File JSON đã sẵn sàng tải xuống!", "success", auto_close=True)

                    elif "PDF" in export_format:
                        # PDF Export
                        try:
                            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                            from reportlab.lib import colors
                            from reportlab.lib.pagesizes import A4, landscape
                            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                            from reportlab.pdfbase import pdfmetrics
                            from reportlab.pdfbase.ttfonts import TTFont
                            from reportlab.lib.units import inch
                            import os
                            
                            # Thử đăng ký font tiếng Việt
                            font_paths = [
                                '/System/Library/Fonts/Arial.ttf',  # macOS
                                'C:\\Windows\\Fonts\\arial.ttf',    # Windows  
                                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
                                '/usr/share/fonts/TTF/DejaVuSans.ttf'  # Linux alt
                            ]
                            
                            vietnamese_font = 'Helvetica'  # Default fallback
                            for font_path in font_paths:
                                try:
                                    if os.path.exists(font_path):
                                        pdfmetrics.registerFont(TTFont('VietnameseFont', font_path))
                                        vietnamese_font = 'VietnameseFont'
                                        break
                                except:
                                    continue
                            
                            def create_pdf():
                                pdf_output = BytesIO()
                                doc = SimpleDocTemplate(
                                    pdf_output,
                                    pagesize=landscape(A4),
                                    topMargin=0.75*inch,
                                    bottomMargin=0.75*inch,
                                    leftMargin=0.5*inch,
                                    rightMargin=0.5*inch
                                )
                                
                                # Custom styles
                                styles = getSampleStyleSheet()
                                title_style = ParagraphStyle(
                                    'VietTitle',
                                    parent=styles['Title'],
                                    fontName=vietnamese_font,
                                    fontSize=18,
                                    textColor=colors.HexColor('#1E40AF'),
                                    alignment=1,
                                    spaceAfter=20
                                )
                                
                                normal_style = ParagraphStyle(
                                    'VietNormal',
                                    parent=styles['Normal'],
                                    fontName=vietnamese_font,
                                    fontSize=10,
                                    textColor=colors.HexColor('#374151')
                                )
                                
                                story = []
                                
                                # Title
                                title = Paragraph("📋 BÁO CÁO ACTIVITY LOGS", title_style)
                                subtitle = Paragraph(f"Tạo lúc: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}", normal_style)
                                story.extend([title, Spacer(1, 20), subtitle, Spacer(1, 30)])
                                
                                # Thống kê
                                stats_title = Paragraph("📊 THỐNG KÊ TỔNG QUAN", normal_style)
                                story.append(stats_title)
                                story.append(Spacer(1, 10))
                                
                                stats_data = [
                                    ['Thông tin', 'Giá trị'],
                                    ['Tổng số logs', f"{len(df_logs):,}"],
                                    ['Số loại hoạt động', f"{df_logs['action'].nunique() if 'action' in df_logs.columns else 0}"],
                                    ['Số người dùng', f"{df_logs['user_id'].nunique() if 'user_id' in df_logs.columns else 0}"],
                                    ['Thời gian đầu', df_logs['formatted_date'].min() if 'formatted_date' in df_logs.columns else 'N/A'],
                                    ['Thời gian cuối', df_logs['formatted_date'].max() if 'formatted_date' in df_logs.columns else 'N/A']
                                ]
                                
                                stats_table = Table(stats_data, colWidths=[2.5*inch, 2*inch])
                                stats_table.setStyle(TableStyle([
                                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E40AF')),
                                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                                    ('FONTNAME', (0,0), (-1,0), vietnamese_font),
                                    ('FONTSIZE', (0,0), (-1,0), 11),
                                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                                    
                                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
                                    ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#374151')),
                                    ('FONTNAME', (0,1), (-1,-1), vietnamese_font),
                                    ('FONTSIZE', (0,1), (-1,-1), 9),
                                    ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E5E7EB')),
                                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
                                    ('TOPPADDING', (0,1), (-1,-1), 8),
                                    ('BOTTOMPADDING', (0,1), (-1,-1), 8),
                                ]))
                                
                                story.append(stats_table)
                                story.append(Spacer(1, 30))
                                
                                # Bảng dữ liệu chính
                                data_title = Paragraph("📋 CHI TIẾT LOGS (100 bản ghi gần nhất)", normal_style)
                                story.append(data_title)
                                story.append(Spacer(1, 15))
                                
                                # Chọn columns để hiển thị
                                display_cols = ['id', 'action', 'user_id', 'formatted_date']
                                available_cols = [col for col in display_cols if col in df_logs.columns]
                                
                                if 'details' in df_logs.columns and len(available_cols) < 5:
                                    available_cols.append('details')
                                
                                df_display = df_logs[available_cols].head(50)
                                
                                # Tạo headers
                                headers = []
                                col_mappings = {
                                    'id': 'ID',
                                    'action': 'Hoạt động', 
                                    'user_id': 'User ID',
                                    'details': 'Chi tiết',
                                    'formatted_date': 'Thời gian',
                                    'ip_address': 'IP'
                                }
                                
                                for col in available_cols:
                                    headers.append(col_mappings.get(col, col.title()))
                                
                                table_data = [headers]
                                
                                # Thêm dữ liệu
                                for _, row in df_display.iterrows():
                                    row_data = []
                                    for col in available_cols:
                                        cell_value = str(row[col]) if pd.notna(row[col]) else ''
                                        if len(cell_value) > 40:
                                            cell_value = cell_value[:37] + "..."
                                        row_data.append(cell_value)
                                    table_data.append(row_data)
                                
                                # Column widths
                                available_width = 10 * inch
                                col_count = len(available_cols)
                                base_width = available_width / col_count
                                
                                col_widths = []
                                for i, col in enumerate(available_cols):
                                    if col == 'id':
                                        col_widths.append(base_width * 0.6)
                                    elif col == 'details':
                                        col_widths.append(base_width * 1.5)
                                    elif col == 'formatted_date':
                                        col_widths.append(base_width * 1.2)
                                    else:
                                        col_widths.append(base_width)
                                
                                total_width = sum(col_widths)
                                col_widths = [w * available_width / total_width for w in col_widths]
                                
                                main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                                main_table.setStyle(TableStyle([
                                    # Header
                                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E40AF')),
                                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                                    ('FONTNAME', (0,0), (-1,0), vietnamese_font),
                                    ('FONTSIZE', (0,0), (-1,0), 9),
                                    ('BOTTOMPADDING', (0,0), (-1,0), 10),
                                    
                                    # Data
                                    ('BACKGROUND', (0,1), (-1,-1), colors.white),
                                    ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#374151')),
                                    ('FONTNAME', (0,1), (-1,-1), vietnamese_font),
                                    ('FONTSIZE', (0,1), (-1,-1), 8),
                                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
                                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
                                    ('TOPPADDING', (0,1), (-1,-1), 6),
                                    ('BOTTOMPADDING', (0,1), (-1,-1), 6),
                                    ('LEFTPADDING', (0,0), (-1,-1), 6),
                                    ('RIGHTPADDING', (0,0), (-1,-1), 6),
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                ]))
                                
                                story.append(main_table)
                                
                                # Footer
                                story.append(Spacer(1, 20))
                                footer = Paragraph(
                                    f"Báo cáo tự động - AI Healthcare Admin - {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}",
                                    ParagraphStyle(
                                        'Footer',
                                        parent=styles['Normal'],
                                        fontName=vietnamese_font,
                                        fontSize=8,
                                        textColor=colors.HexColor('#9CA3AF'),
                                        alignment=1
                                    )
                                )
                                story.append(footer)
                                
                                doc.build(story)
                                return pdf_output.getvalue()
                            
                            pdf_data = create_pdf()
                            st.download_button(
                                "📄 Tải file PDF",
                                pdf_data,
                                f"activity_logs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.pdf",
                                "application/pdf",
                                help="PDF chuyên nghiệp với font tiếng Việt và layout đẹp",
                                use_container_width=True
                            )
                            show_notification("File PDF đã sẵn sàng tải xuống!", "success", auto_close=True)
                            
                        except ImportError:
                            st.error("⚠️ Thiếu thư viện reportlab. Vui lòng cài đặt: `pip install reportlab`")
                        except Exception as e:
                            st.error(f"❌ Lỗi tạo PDF: {str(e)}")

                except Exception as e:
                    show_notification(f"Lỗi khi xuất file: {str(e)}", "error")

            # Thông báo hướng dẫn - Cập nhật cho dropdown
            st.markdown("""
            <div style="background: #EFF6FF; border-left: 4px solid #3B82F6; padding: 1rem; margin: 1rem 0; border-radius: 0 8px 8px 0;">
                <h5 style="color: #1E40AF; margin: 0 0 0.5rem 0;">💡 Hướng dẫn xuất file:</h5>
                <ul style="color: #374151; margin: 0; padding-left: 1.5rem;">
                    <li><strong>CSV:</strong> Có BOM UTF-8 - Excel sẽ tự động hiển thị tiếng Việt đúng</li>
                    <li><strong>Excel:</strong> Định dạng header đẹp, auto-resize, có sheet thống kê riêng</li>
                    <li><strong>JSON:</strong> Giữ nguyên ký tự tiếng Việt, format đẹp với indent</li>
                    <li><strong>PDF:</strong> Tự động detect font hệ thống, layout ngang, có thống kê</li>
                </ul>
                <div style="margin-top: 0.75rem; padding: 0.5rem; background: #DBEAFE; border-radius: 4px;">
                    <strong>🔥 Mới:</strong> Chọn định dạng từ dropdown và nhấn "Tải xuống" - Giao diện gọn gàng hơn!
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Additional Export Options
            st.markdown("---")
            
            # Bulk Export Section
            with st.expander("🎯 Xuất dữ liệu nâng cao", expanded=False):
                st.markdown("""
                <div style="background: #f8fafc; padding: 1rem; border-radius: 8px; margin: 1rem 0;">
                    <h5 style="color: #1e40af; margin-bottom: 0.5rem;">📊 Tùy chọn xuất dữ liệu</h5>
                </div>
                """, unsafe_allow_html=True)
                
                export_col1, export_col2 = st.columns(2)
                
                with export_col1:
                    # Date range filter
                    st.write("📅 **Lọc theo thời gian:**")
                    if 'created_at' in df_logs.columns:
                        date_from = st.date_input("Từ ngày:", value=pd.Timestamp.now().date() - pd.Timedelta(days=30))
                        date_to = st.date_input("Đến ngày:", value=pd.Timestamp.now().date())
                
                with export_col2:
                    # Column selection
                    st.write("📋 **Chọn cột xuất:**")
                    selected_columns = st.multiselect(
                        "Columns:",
                        options=df_logs.columns.tolist(),
                        default=df_logs.columns.tolist()[:5]  # First 5 columns by default
                    )
                
                # Custom export button
                if st.button("🚀 Xuất dữ liệu tùy chỉnh", key="custom_export"):
                    if selected_columns:
                        filtered_data = df_logs[selected_columns]
                        
                        # Apply date filter if available
                        if 'created_at' in df_logs.columns:
                            filtered_data = filtered_data[
                                (pd.to_datetime(df_logs['created_at']).dt.date >= date_from) &
                                (pd.to_datetime(df_logs['created_at']).dt.date <= date_to)
                            ]
                        
                        # Tạo CSV với UTF-8 BOM
                        csv_buffer = io.StringIO()
                        filtered_data.to_csv(csv_buffer, index=False)
                        csv_content = csv_buffer.getvalue()
                        custom_csv = ('\ufeff' + csv_content).encode('utf-8')
                        
                        st.download_button(
                            "⬇️ Tải CSV tùy chỉnh",
                            custom_csv,
                            f"custom_logs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                            "text/csv",
                            help="CSV tùy chỉnh với UTF-8 BOM"
                        )
                        show_notification("Xuất dữ liệu tùy chỉnh thành công!", "success", auto_close=True)
                    else:
                        show_notification("Vui lòng chọn ít nhất một cột để xuất!", "warning")

        else:
            show_notification("Không có logs - Chưa có hoạt động nào được ghi lại trong hệ thống.", "info")
    except Exception as e:
        show_notification(f"Lỗi khi tải logs: {e}", "error")
    
    st.markdown('</div>', unsafe_allow_html=True)
# ==========================
# 🔹 Footer với thông tin hệ thống
# ==========================
st.markdown("---")
st.markdown(f"""
<div style="text-align: center; padding: 2rem 0; background: white; border-radius: 15px; margin-top: 2rem; box-shadow: 0 4px 20px rgba(59, 130, 246, 0.1);">
    <h4 style="color: #1e40af; margin-bottom: 1rem;">🛠️ Admin Control Panel</h4>
    <p style="color: #64748b; margin-bottom: 0.5rem;">
        <strong>Hệ thống quản trị AI Healthcare Platform</strong>
    </p>
    <p style="color: #64748b; font-size: 0.9rem;">
        Phiên bản: 2.1.0 | Cập nhật: {pd.Timestamp.now().strftime('%d/%m/%Y')} | 
        Người dùng: <strong>{st.session_state.user.get('username', 'Admin')}</strong>
    </p>
    <div style="margin-top: 1rem; color: #64748b; font-size: 0.8rem;">
        💡 <em>Với quyền Admin, bạn có toàn quyền kiểm soát và giám sát hệ thống</em>
    </div>
    <div style="margin-top: 0.5rem; color: #ef4444; font-size: 0.8rem; font-weight: 600;">
        ⚠️ <em>Hãy thận trọng khi sử dụng các chức năng khóa/xóa tài khoản!</em>
    </div>
</div>
""", unsafe_allow_html=True)
